"""Writes submittals to the shared log in a banded-table tracker format:
one Excel Table per project sheet, the Submittal cell hyperlinked to the
dropped PDF, and revision handling — a re-submitted submittal updates its
existing row (approvals reset) while the superseded version moves to the
'Revision History' sheet with a link to the old PDF."""
import difflib
import os
import re
from datetime import date, datetime
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Column layout for the per-project submittal tracker
# (with dedicated Submittal # and Rev. columns)
COLUMNS = [
    "#", "Submittal #", "Rev.", "Trade", "Submittal", "Contractor",
    "Initial Submission", "Structural Approval", "MEP Approval",
    "Interior Approval", "Consultant Approval", "Architect Approval",
    "Approved", "Approval Date", "Material Release", "Lead Time Weeks",
    "ETA", "Notes",
]
WIDTHS = [5, 15, 6, 16, 61, 24, 18, 21, 14, 17, 20, 18, 11, 16, 17, 18, 12, 50]
APPROVAL_COLS = ["Structural Approval", "MEP Approval", "Interior Approval",
                 "Consultant Approval", "Architect Approval", "Approval Date"]

HISTORY_SHEET = "Revision History"
HISTORY_COLUMNS = ["Project", "Submittal #", "Rev.", "Trade", "Submittal",
                   "Contractor", "Initial Submission", "Approved", "Notes",
                   "Superseded On"]
HISTORY_WIDTHS = [24, 15, 6, 16, 55, 24, 18, 11, 45, 14]

DATE_FMT = "mm-dd-yy"
LINK_FONT = Font(name="Aptos Narrow", size=11, color="0563C1", underline="single")
INBOX_SUBDIR = "Submittals Inbox"   # PDFs live here, next to the log workbook


_SUFFIXES = r"\b(road|rd|street|st|avenue|ave|av|boulevard|blvd|drive|dr|lane|ln|place|pl)\b"

def _norm(s):
    s = str(s or "").lower()
    s = re.sub(_SUFFIXES, "", s)            # "main street" == "main st"
    return re.sub(r"[^a-z0-9]", "", s)


def _table_name(title):
    base = "T_" + re.sub(r"[^A-Za-z0-9_]", "", title)
    return base[:255] or "T_Sheet"


def _ensure_table(ws, ncols=None):
    """Create/resize this sheet's banded table to cover all rows."""
    ncols = ncols or len(COLUMNS)
    ref = f"A1:{get_column_letter(ncols)}{max(ws.max_row, 2)}"
    if ws.tables:
        for t in ws.tables.values():
            t.ref = ref
        return
    t = Table(displayName=_table_name(ws.title), ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",
                                      showRowStripes=True)
    ws.add_table(t)


def _new_sheet(wb, title, columns, widths):
    ws = wb.create_sheet(title[:31])
    ws.append(columns)
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    return ws


def _project_sheet(wb, project):
    """Find the sheet for this project (fuzzy) or create one."""
    want = _norm(project) or "unassigned"
    for ws in wb.worksheets:
        if ws.title == HISTORY_SHEET:
            continue
        n = _norm(ws.title)
        if not n:
            continue
        if n in want or want in n or difflib.SequenceMatcher(None, n, want).ratio() > 0.8:
            return ws
    return _new_sheet(wb, project or "Unassigned", COLUMNS, WIDTHS)


def _parse_date(s):
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(s), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _lead_time_weeks(s):
    """'6 weeks' -> 6; '30 days' -> 4; passthrough otherwise."""
    if not s:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*(week|wk)", str(s), re.I)
    if m:
        return float(m.group(1))
    m = re.search(r"(\d+(?:\.\d+)?)\s*day", str(s), re.I)
    if m:
        return round(float(m.group(1)) / 7, 1)
    return str(s)


def _apply_approval(cell, data):
    """If the PDF carried a reviewer disposition stamp, record it VERBATIM:
    the Approved column says what the stamp says ("Approved", "Approved as
    Noted", "Revise & Resubmit", ...), the matching discipline column gets
    the stamp text+date, and an approving stamp sets Approval Date."""
    status = str(data.get("approval_status") or "").strip()
    if not status:
        return None
    app_date = _parse_date(data.get("approval_date"))
    disc = str(data.get("approval_discipline") or "").strip().lower()
    for col in COLUMNS:
        if col.lower() == f"{disc} approval":
            cell(col).value = status + (f" {app_date:%m/%d/%Y}" if app_date else "")
            break
    cell("Approved").value = status
    if app_date and ("approved" in status.lower() or "no exception" in status.lower()):
        c = cell("Approval Date")
        c.value = app_date
        c.number_format = DATE_FMT
    return status


# ------------------------------------------------------------- revisions

def _norm_num(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def _norm_rev(v):
    """'Rev2' -> 2, '00' -> 0, 5 -> 5; None/'' -> None; 'B' stays 'B'."""
    if v in (None, ""):
        return None
    m = re.search(r"\d+", str(v))
    return int(m.group()) if m else str(v)


def _base_num(s):
    """'260000-008-3' -> '260000008': the submittal id with a trailing
    1-2 digit revision segment stripped, for rev-to-rev matching."""
    m = re.match(r"^(.*?)[-_. ]\d{1,2}$", str(s or "").strip())
    return _norm_num(m.group(1)) if m else _norm_num(s)


def _find_revision_row(ws, data):
    """Row index of an earlier revision of this submittal, or None."""
    new_num = _norm_num(data.get("submittal_number"))
    new_base = _base_num(data.get("submittal_number"))
    new_title = str(data.get("title") or "").lower()
    num_col = COLUMNS.index("Submittal #") + 1
    title_col = COLUMNS.index("Submittal") + 1
    for r in range(2, ws.max_row + 1):
        old_raw = ws.cell(r, num_col).value
        old_num = _norm_num(old_raw)
        old_base = _base_num(old_raw)
        old_title = str(ws.cell(r, title_col).value or "").lower()
        if new_num and old_num and new_num == old_num:
            return r
        if (len(new_base) >= 3 and new_base == old_base):
            return r
        if (new_title and old_title and
                difflib.SequenceMatcher(None, new_title, old_title).ratio() >= 0.87):
            return r
    return None


def _history_sheet(wb):
    if HISTORY_SHEET in wb.sheetnames:
        return wb[HISTORY_SHEET]
    return _new_sheet(wb, HISTORY_SHEET, HISTORY_COLUMNS, HISTORY_WIDTHS)


def _archive_revision(wb, ws, r):
    """Copy row r of a project sheet onto the Revision History sheet."""
    h = _history_sheet(wb)
    cell = lambda name: ws.cell(r, COLUMNS.index(name) + 1)
    h.append([ws.title, cell("Submittal #").value, cell("Rev.").value,
              cell("Trade").value, cell("Submittal").value,
              cell("Contractor").value, cell("Initial Submission").value,
              cell("Approved").value, cell("Notes").value, date.today()])
    hr = h.max_row
    h.cell(hr, HISTORY_COLUMNS.index("Initial Submission") + 1).number_format = DATE_FMT
    h.cell(hr, HISTORY_COLUMNS.index("Superseded On") + 1).number_format = DATE_FMT
    old_link = cell("Submittal").hyperlink
    if old_link:
        c = h.cell(hr, HISTORY_COLUMNS.index("Submittal") + 1)
        c.hyperlink = old_link.target
        c.font = LINK_FONT
    _ensure_table(h, len(HISTORY_COLUMNS))


# ----------------------------------------------------------------- write

def _build_notes(data):
    notes = []
    if data.get("spec_section"):
        notes.append(str(data["spec_section"]))
    if data.get("description"):
        notes.append(str(data["description"]))
    return notes


def append_submittal(xlsx_path: str, data: dict, pdf_filename: str = None,
                     link_url: str = None) -> str:
    """Log one extracted submittal: append a new row, or — when it's a new
    revision of an existing row — update that row in place and archive the
    old version. Returns a description of what was written.

    link_url: explicit hyperlink target (e.g. a Dropbox share URL from the
    cloud watcher); falls back to a relative link into the inbox folder."""
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    ws = _project_sheet(wb, data.get("project"))
    sub_date = _parse_date(data.get("submittal_date")) or date.today()
    rev = _norm_rev(data.get("revision"))
    link = link_url or (quote(f"{INBOX_SUBDIR}/{pdf_filename}", safe="/")
                        if pdf_filename else None)

    prev = _find_revision_row(ws, data)
    if prev:
        # ---- new revision of an existing submittal: update in place
        _archive_revision(wb, ws, prev)
        cell = lambda name: ws.cell(prev, COLUMNS.index(name) + 1)
        old_rev = _norm_rev(cell("Rev.").value)
        if isinstance(old_rev, int) and not (isinstance(rev, int) and rev > old_rev):
            rev = old_rev + 1          # extractor gave nothing newer: bump
        elif rev is None:
            rev = "resub"
        notes = _build_notes(data)
        notes.append(f"Rev. {rev} received {date.today():%m/%d/%Y} — "
                     f"approvals reset; prior version on '{HISTORY_SHEET}'")

        if data.get("submittal_number"):
            cell("Submittal #").value = data["submittal_number"]
        cell("Rev.").value = rev
        if data.get("trade") or data.get("type"):
            cell("Trade").value = data.get("trade") or data.get("type")
        if data.get("title"):
            cell("Submittal").value = data["title"]
        if data.get("responsible_contractor"):
            cell("Contractor").value = data["responsible_contractor"]
        for name in APPROVAL_COLS:
            cell(name).value = None
        cell("Approved").value = "NO"
        _apply_approval(cell, data)
        if data.get("lead_time"):
            cell("Lead Time Weeks").value = _lead_time_weeks(data["lead_time"])
        cell("Notes").value = " · ".join(notes)
        if link:
            c = cell("Submittal")
            c.hyperlink = link
            c.font = LINK_FONT
        wb.save(xlsx_path)
        return f"{ws.title}' row {prev} UPDATED to Rev. {rev} (old rev archived)"

    # ---- brand-new submittal: append
    seq = max(ws.max_row - 1, 0) + 1
    row = {
        "#": seq,
        "Submittal #": data.get("submittal_number"),
        "Rev.": rev if rev is not None else 0,
        "Trade": data.get("trade") or data.get("type"),
        "Submittal": data.get("title") or pdf_filename,
        "Contractor": data.get("responsible_contractor"),
        "Initial Submission": sub_date,
        "Approved": "NO",   # flipped to YES manually once it clears review
        "Lead Time Weeks": _lead_time_weeks(data.get("lead_time")),
        "Notes": " · ".join(_build_notes(data)) or None,
    }
    ws.append([row.get(c) for c in COLUMNS])

    r = ws.max_row
    ws.cell(r, COLUMNS.index("Initial Submission") + 1).number_format = DATE_FMT
    cell = lambda name: ws.cell(r, COLUMNS.index(name) + 1)
    _apply_approval(cell, data)
    if link:
        link_cell = ws.cell(r, COLUMNS.index("Submittal") + 1)
        # relative to the workbook, which sits next to the inbox folder;
        # spaces must be %-encoded or Excel can't resolve the target
        link_cell.hyperlink = link
        link_cell.font = LINK_FONT

    _ensure_table(ws)
    wb.save(xlsx_path)
    return ws.title
